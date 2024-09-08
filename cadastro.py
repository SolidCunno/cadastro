# Aqui decidi usar tkinter para uma GUI pois todo o cadastro era feito diretamente na planilha
# openpyxl para manipular o arquivo .xlsx

import tkinter as tk
from tkinter import messagebox
from openpyxl import Workbook, load_workbook
import os


class Pessoa:
    def __init__(self, nome, sobrenome, cpf, data_nascimento, rg, numero, email, cep, endereco, data_cadastro, responsavel, atividade):
        self.nome = nome
        self.sobrenome = sobrenome
        self.cpf = cpf
        self.data_nascimento = data_nascimento
        self.rg = rg
        self.numero = numero
        self.email = email
        self.cep = cep
        self.endereco = endereco
        self.data_cadastro = data_cadastro
        self.responsavel = responsavel
        self.atividade = atividade


def salvar_dados(pessoa):
    if os.path.exists('Cadastro_de_Clientes_geral.xlsx'):
        workbook = load_workbook('Cadastro_de_Clientes_geral.xlsx')
        sheet = workbook.active
    else:
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["Nome", "Sobrenome", "CPF", "Data de Nascimento", "RG", "Número", "Email", "CEP", "Endereço", "Data de Cadastro", "Responsável", "Atividade"])

    sheet.append([pessoa.nome, pessoa.sobrenome, pessoa.cpf, pessoa.data_nascimento, pessoa.rg, pessoa.numero, pessoa.email, pessoa.cep, pessoa.endereco, pessoa.data_cadastro, pessoa.responsavel, pessoa.atividade])
    workbook.save('Cadastro_de_Clientes_geral.xlsx')


def buscar_pessoa_por_nome(nome):
    if os.path.exists('Cadastro_de_Clientes_geral.xlsx'):
        workbook = load_workbook('Cadastro_de_Clientes_geral.xlsx')
        sheet = workbook.active
        for row in sheet.iter_rows(values_only=True):
            if row[0] == nome:
                return Pessoa(*row)
    return None


def buscar_pessoa_por_cpf(cpf):
    if os.path.exists('Cadastro_de_Clientes_geral.xlsx'):
        workbook = load_workbook('Cadastro_de_Clientes_geral.xlsx')
        sheet = workbook.active
        for row in sheet.iter_rows(values_only=True):
            if row[2] == cpf:
                return Pessoa(*row)
    return None


def limpar_campos():
    nome_entry.delete(0, tk.END)
    sobrenome_entry.delete(0, tk.END)
    cpf_entry.delete(0, tk.END)
    data_nascimento_entry.delete(0, tk.END)
    rg_entry.delete(0, tk.END)
    numero_entry.delete(0, tk.END)
    email_entry.delete(0, tk.END)
    cep_entry.delete(0, tk.END)
    endereco_entry.delete(0, tk.END)
    data_cadastro_entry.delete(0, tk.END)
    responsavel_entry.delete(0, tk.END)
    atividade_entry.delete(0, tk.END)


def salvar():
    nome = nome_entry.get()
    sobrenome = sobrenome_entry.get()
    cpf = cpf_entry.get()
    data_nascimento = data_nascimento_entry.get()
    rg = rg_entry.get()
    numero = numero_entry.get()
    email = email_entry.get()
    cep = cep_entry.get()
    endereco = endereco_entry.get()
    data_cadastro = data_cadastro_entry.get()
    responsavel = responsavel_entry.get()
    atividade = atividade_entry.get()

    if nome and sobrenome and cpf and data_nascimento and rg and numero and email and cep and endereco and data_cadastro and responsavel and atividade:
        pessoa = Pessoa(nome, sobrenome, cpf, data_nascimento, rg, numero, email, cep, endereco, data_cadastro, responsavel, atividade)
        salvar_dados(pessoa)
        messagebox.showinfo("Sucesso", "Cadastro concluído com sucesso!")
        limpar_campos()
    else:
        messagebox.showwarning("Atenção", "Todos os campos são obrigatórios!")


def buscar_por_nome():
    nome = busca_nome_entry.get()
    pessoa = buscar_pessoa_por_nome(nome)
    if pessoa:
        resultado_text.delete(1.0, tk.END)
        resultado_text.insert(tk.END, f"Nome: {pessoa.nome}\nSobrenome: {pessoa.sobrenome}\nCPF: {pessoa.cpf}\nData de Nascimento: {pessoa.data_nascimento}\nRG: {pessoa.rg}\nNúmero: {pessoa.numero}\nEmail: {pessoa.email}\nCEP: {pessoa.cep}\nEndereço: {pessoa.endereco}\nData de Cadastro: {pessoa.data_cadastro}\nResponsável: {pessoa.responsavel}\nAtividade: {pessoa.atividade}")
    else:
        messagebox.showinfo("Resultado", "Usuário não encontrado.")


def buscar_por_cpf():
    cpf = busca_cpf_entry.get()
    pessoa = buscar_pessoa_por_cpf(cpf)
    if pessoa:
        resultado_text.delete(1.0, tk.END)
        resultado_text.insert(tk.END, f"Nome: {pessoa.nome}\nSobrenome: {pessoa.sobrenome}\nCPF: {pessoa.cpf}\nData de Nascimento: {pessoa.data_nascimento}\nRG: {pessoa.rg}\nNúmero: {pessoa.numero}\nEmail: {pessoa.email}\nCEP: {pessoa.cep}\nEndereço: {pessoa.endereco}\nData de Cadastro: {pessoa.data_cadastro}\nResponsável: {pessoa.responsavel}\nAtividade: {pessoa.atividade}")
    else:
        messagebox.showinfo("Resultado", "Usuário não encontrado.")

# GUI
root = tk.Tk()
root.title("Projeto Alto do Coqueirinho Em Movimento")  # Adicionando o título da janela

tk.Label(root, text="Nome:").grid(row=0, column=0)
nome_entry = tk.Entry(root)
nome_entry.grid(row=0, column=1)

tk.Label(root, text="Sobrenome:").grid(row=1, column=0)
sobrenome_entry = tk.Entry(root)
sobrenome_entry.grid(row=1, column=1)

tk.Label(root, text="CPF:").grid(row=2, column=0)
cpf_entry = tk.Entry(root)
cpf_entry.grid(row=2, column=1)

tk.Label(root, text="Data de Nascimento:").grid(row=3, column=0)
data_nascimento_entry = tk.Entry(root)
data_nascimento_entry.grid(row=3, column=1)

tk.Label(root, text="RG:").grid(row=4, column=0)
rg_entry = tk.Entry(root)
rg_entry.grid(row=4, column=1)

tk.Label(root, text="Número:").grid(row=5, column=0)
numero_entry = tk.Entry(root)
numero_entry.grid(row=5, column=1)

tk.Label(root, text="Email:").grid(row=6, column=0)
email_entry = tk.Entry(root)
email_entry.grid(row=6, column=1)

tk.Label(root, text="CEP:").grid(row=7, column=0)
cep_entry = tk.Entry(root)
cep_entry.grid(row=7, column=1)

tk.Label(root, text="Endereço:").grid(row=8, column=0)
endereco_entry = tk.Entry(root)
endereco_entry.grid(row=8, column=1)

tk.Label(root, text="Data de Cadastro:").grid(row=9, column=0)
data_cadastro_entry = tk.Entry(root)
data_cadastro_entry.grid(row=9, column=1)

tk.Label(root, text="Responsável:").grid(row=10, column=0)
responsavel_entry = tk.Entry(root)
responsavel_entry.grid(row=10, column=1)

tk.Label(root, text="Atividade:").grid(row=11, column=0)
atividade_entry = tk.Entry(root)
atividade_entry.grid(row=11, column=1)

tk.Button(root, text="Salvar", command=salvar).grid(row=12, column=0, columnspan=2)

tk.Label(root, text="Buscar por Nome:").grid(row=13, column=0)
busca_nome_entry = tk.Entry(root)
busca_nome_entry.grid(row=13, column=1)

tk.Button(root, text="Buscar", command=buscar_por_nome).grid(row=14, column=0, columnspan=2)

tk.Label(root, text="Buscar por CPF:").grid(row=15, column=0)
busca_cpf_entry = tk.Entry(root)
busca_cpf_entry.grid(row=15, column=1)

tk.Button(root, text="Buscar", command=buscar_por_cpf).grid(row=16, column=0, columnspan=2)

resultado_text = tk.Text(root, height=10, width=50)
resultado_text.grid(row=17, column=0, columnspan=2)

root.mainloop()
